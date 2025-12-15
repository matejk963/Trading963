"""
Create Excel Report with CoT Positioning Analysis
- Main dashboard with z-scores by category
- Configurable lookback period
- Individual sheets with charts for each contract
"""
import sys
sys.path.insert(0, '/mnt/c/Users/krajcovic/Documents/GitHub/Trading963')

from src.data_fetchers.cftc_fetcher import CFTCLegacyFetcher
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

import argparse

# Parse command line arguments
parser = argparse.ArgumentParser(description='Generate CoT Positioning Excel Report')
parser.add_argument('--lookback', type=int, default=3, help='Lookback period in years (default: 3)')
args = parser.parse_args()

lookback_years = args.lookback

print("="*80)
print("CREATING EXCEL COT POSITIONING REPORT")
print(f"Lookback Period: {lookback_years} years")
print("="*80)

# Load data
fetcher = CFTCLegacyFetcher()
all_data = []
years = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]

print("\nLoading data...")
for year in years:
    file_path = f"data/cftc/legacy_long_format/{year}/annual.txt"
    df = fetcher.load_legacy_data(file_path)
    if df is not None:
        all_data.append(df)
        print(f"  ✓ {year}")

combined_df = pd.concat(all_data, ignore_index=True)
combined_df.columns = combined_df.columns.str.replace(' ', '_').str.replace('(', '').str.replace(')', '')
combined_df['Date'] = pd.to_datetime(combined_df['As_of_Date_in_Form_YYYY-MM-DD'])

# Define all contracts with their codes
contracts = {
    # Currencies
    'EURO FX - CHICAGO MERCANTILE EXCHANGE': {'code': '6E', 'category': 'Currencies', 'name': 'Euro FX'},
    'JAPANESE YEN - CHICAGO MERCANTILE EXCHANGE': {'code': '6J', 'category': 'Currencies', 'name': 'Japanese Yen'},
    'BRITISH POUND STERLING - CHICAGO MERCANTILE EXCHANGE': {'code': '6B', 'category': 'Currencies', 'name': 'British Pound'},
    'CANADIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE': {'code': '6C', 'category': 'Currencies', 'name': 'Canadian Dollar'},
    'SWISS FRANC - CHICAGO MERCANTILE EXCHANGE': {'code': '6S', 'category': 'Currencies', 'name': 'Swiss Franc'},
    'AUSTRALIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE': {'code': '6A', 'category': 'Currencies', 'name': 'Australian Dollar'},
    'NEW ZEALAND DOLLAR - CHICAGO MERCANTILE EXCHANGE': {'code': '6N', 'category': 'Currencies', 'name': 'New Zealand Dollar'},

    # Indices
    'S&P 500 Consolidated': {'code': 'ES', 'category': 'Indices', 'name': 'S&P 500 E-Mini'},
    'NASDAQ STOCK MARKET MINI INDEX - CHICAGO MERCANTILE EXCHANGE': {'code': 'NQ', 'category': 'Indices', 'name': 'Nasdaq 100'},
    'DOW JONES INDUSTRIAL AVERAGE MINI - CHICAGO BOARD OF TRADE': {'code': 'YM', 'category': 'Indices', 'name': 'Dow Jones'},
    'RUSSELL 2000 MINI INDEX FUTURE - ICE FUTURES U.S.': {'code': 'RTY', 'category': 'Indices', 'name': 'Russell 2000'},
    'VIX FUTURES - CHICAGO BOARD OPTIONS EXCHANGE': {'code': 'VX', 'category': 'Indices', 'name': 'VIX'},

    # Energy
    'CRUDE OIL, LIGHT SWEET - NEW YORK MERCANTILE EXCHANGE': {'code': 'CL', 'category': 'Energy', 'name': 'Crude Oil WTI'},
    'BRENT LAST DAY FINANCIAL - NEW YORK MERCANTILE EXCHANGE': {'code': 'BZ', 'category': 'Energy', 'name': 'Brent Crude'},
    'NATURAL GAS - NEW YORK MERCANTILE EXCHANGE': {'code': 'NG', 'category': 'Energy', 'name': 'Natural Gas'},
    'GASOLINE RBOB - NEW YORK MERCANTILE EXCHANGE': {'code': 'RB', 'category': 'Energy', 'name': 'RBOB Gasoline'},
    'NY HARBOR ULSD - NEW YORK MERCANTILE EXCHANGE': {'code': 'HO', 'category': 'Energy', 'name': 'Heating Oil'},

    # Metals
    'GOLD - COMMODITY EXCHANGE INC.': {'code': 'GC', 'category': 'Metals', 'name': 'Gold'},
    'SILVER - COMMODITY EXCHANGE INC.': {'code': 'SI', 'category': 'Metals', 'name': 'Silver'},
    'COPPER - COMMODITY EXCHANGE INC.': {'code': 'HG', 'category': 'Metals', 'name': 'Copper'},
    'PLATINUM - NEW YORK MERCANTILE EXCHANGE': {'code': 'PL', 'category': 'Metals', 'name': 'Platinum'},
    'PALLADIUM - NEW YORK MERCANTILE EXCHANGE': {'code': 'PA', 'category': 'Metals', 'name': 'Palladium'},

    # Grains
    'CORN - CHICAGO BOARD OF TRADE': {'code': 'ZC', 'category': 'Grains', 'name': 'Corn'},
    'WHEAT - CHICAGO BOARD OF TRADE': {'code': 'ZW', 'category': 'Grains', 'name': 'Wheat'},
    'SOYBEANS - CHICAGO BOARD OF TRADE': {'code': 'ZS', 'category': 'Grains', 'name': 'Soybeans'},
    'SOYBEAN MEAL - CHICAGO BOARD OF TRADE': {'code': 'ZM', 'category': 'Grains', 'name': 'Soybean Meal'},
    'SOYBEAN OIL - CHICAGO BOARD OF TRADE': {'code': 'ZL', 'category': 'Grains', 'name': 'Soybean Oil'},
    'OATS - CHICAGO BOARD OF TRADE': {'code': 'ZO', 'category': 'Grains', 'name': 'Oats'},
    'ROUGH RICE - CHICAGO BOARD OF TRADE': {'code': 'ZR', 'category': 'Grains', 'name': 'Rough Rice'},

    # Softs
    'COFFEE C - ICE FUTURES U.S.': {'code': 'KC', 'category': 'Softs', 'name': 'Coffee'},
    'SUGAR NO. 11 - ICE FUTURES U.S.': {'code': 'SB', 'category': 'Softs', 'name': 'Sugar'},
    'COCOA - ICE FUTURES U.S.': {'code': 'CC', 'category': 'Softs', 'name': 'Cocoa'},
    'COTTON NO. 2 - ICE FUTURES U.S.': {'code': 'CT', 'category': 'Softs', 'name': 'Cotton'},
    'ORANGE JUICE FROZEN CONCENTRATE - ICE FUTURES U.S.': {'code': 'OJ', 'category': 'Softs', 'name': 'Orange Juice'},
    'LUMBER - CHICAGO MERCANTILE EXCHANGE': {'code': 'LBS', 'category': 'Softs', 'name': 'Lumber'},

    # Meats
    'LIVE CATTLE - CHICAGO MERCANTILE EXCHANGE': {'code': 'LE', 'category': 'Meats', 'name': 'Live Cattle'},
    'FEEDER CATTLE - CHICAGO MERCANTILE EXCHANGE': {'code': 'GF', 'category': 'Meats', 'name': 'Feeder Cattle'},
    'LEAN HOGS - CHICAGO MERCANTILE EXCHANGE': {'code': 'HE', 'category': 'Meats', 'name': 'Lean Hogs'},

    # Bonds
    'TREASURY NOTE 10-YEAR - CHICAGO BOARD OF TRADE': {'code': 'ZN', 'category': 'Bonds', 'name': '10-Year T-Note'},
    'TREASURY NOTE 5-YEAR - CHICAGO BOARD OF TRADE': {'code': 'ZF', 'category': 'Bonds', 'name': '5-Year T-Note'},
    'TREASURY NOTE 2-YEAR - CHICAGO BOARD OF TRADE': {'code': 'ZT', 'category': 'Bonds', 'name': '2-Year T-Note'},
}

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
category_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create Dashboard sheet
print("\nCreating Dashboard...")
dashboard = wb.create_sheet("Dashboard", 0)

# Add title
dashboard['A1'] = "CFTC COMMITMENTS OF TRADERS - POSITIONING DASHBOARD"
dashboard['A1'].font = Font(bold=True, size=14)
dashboard['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

# Add parameter cell
dashboard['A4'] = "LOOKBACK PERIOD (Years):"
dashboard['A4'].font = Font(bold=True)
dashboard['B4'] = lookback_years
dashboard['B4'].font = Font(size=12)
dashboard['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

dashboard['A5'] = "Note: After changing lookback, save & close Excel, then double-click: tools/update_cot_report.bat"
dashboard['A5'].font = Font(italic=True, size=9)

# Headers
row = 7
dashboard[f'A{row}'] = "Code"
dashboard[f'B{row}'] = "Name"
dashboard[f'C{row}'] = "Category"
dashboard[f'D{row}'] = "Current %"
dashboard[f'E{row}'] = "Z-Score"
dashboard[f'F{row}'] = "Percentile"
dashboard[f'G{row}'] = "Status"
dashboard[f'H{row}'] = "Min (3yr)"
dashboard[f'I{row}'] = "Max (3yr)"
dashboard[f'J{row}'] = "Latest Date"

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    cell = dashboard[f'{col}{row}']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Set column widths
dashboard.column_dimensions['A'].width = 8
dashboard.column_dimensions['B'].width = 20
dashboard.column_dimensions['C'].width = 12
dashboard.column_dimensions['D'].width = 12
dashboard.column_dimensions['E'].width = 10
dashboard.column_dimensions['F'].width = 12
dashboard.column_dimensions['G'].width = 15
dashboard.column_dimensions['H'].width = 12
dashboard.column_dimensions['I'].width = 12
dashboard.column_dimensions['J'].width = 12

row += 1

# Process each contract and calculate stats
print("\nProcessing contracts...")
lookback_date = datetime.now() - timedelta(days=lookback_years*365)

dashboard_data = []
contract_data = {}

for contract_name, info in contracts.items():
    # Get contract data
    contract_df = combined_df[combined_df['Market_and_Exchange_Names'] == contract_name].copy()

    # Handle S&P 500 Consolidated
    if len(contract_df) == 0 and 'S&P 500' in contract_name:
        emini_old = combined_df[combined_df['Market_and_Exchange_Names'] == 'E-MINI S&P 500 STOCK INDEX - CHICAGO MERCANTILE EXCHANGE']
        emini_new = combined_df[combined_df['Market_and_Exchange_Names'] == 'E-MINI S&P 500 - CHICAGO MERCANTILE EXCHANGE']
        contract_df = pd.concat([emini_old, emini_new])

    if len(contract_df) == 0:
        print(f"  ⚠ Skipping {info['code']} - no data")
        continue

    # Calculate commercial net position as % of OI
    contract_df['Commercial_Net_Long'] = (
        contract_df['Commercial_Positions-Long_All'] -
        contract_df['Commercial_Positions-Short_All']
    )
    contract_df['Commercial_Net_Pct'] = (
        contract_df['Commercial_Net_Long'] / contract_df['Open_Interest_All']
    ) * 100

    contract_df = contract_df.sort_values('Date')

    # Filter for lookback period
    lookback_df = contract_df[contract_df['Date'] >= lookback_date].copy()

    if len(lookback_df) < 10:
        print(f"  ⚠ Skipping {info['code']} - insufficient data")
        continue

    # Get latest value
    latest = lookback_df.iloc[-1]
    current_pct = latest['Commercial_Net_Pct']
    current_date = latest['Date']

    # Calculate statistics
    hist_pct = lookback_df['Commercial_Net_Pct']
    min_pct = hist_pct.min()
    max_pct = hist_pct.max()
    mean_pct = hist_pct.mean()
    std_pct = hist_pct.std()

    percentile = (hist_pct < current_pct).sum() / len(hist_pct) * 100
    z_score = (current_pct - mean_pct) / std_pct if std_pct > 0 else 0

    # Determine status
    if percentile >= 95:
        status = "EXTREME LONG"
    elif percentile >= 80:
        status = "Strong Long"
    elif percentile <= 5:
        status = "EXTREME SHORT"
    elif percentile <= 20:
        status = "Strong Short"
    else:
        status = "Neutral"

    dashboard_data.append({
        'code': info['code'],
        'name': info['name'],
        'category': info['category'],
        'current_pct': current_pct,
        'z_score': z_score,
        'percentile': percentile,
        'status': status,
        'min_pct': min_pct,
        'max_pct': max_pct,
        'latest_date': current_date
    })

    # Store full contract data for individual sheets
    contract_data[info['code']] = {
        'info': info,
        'df': contract_df[['Date', 'Open_Interest_All', 'Commercial_Net_Pct']].copy(),
        'stats': {
            'mean': mean_pct,
            'std': std_pct,
            'min': min_pct,
            'max': max_pct
        }
    }

    print(f"  ✓ {info['code']} - {info['name']}")

# Sort by category then z-score
dashboard_df = pd.DataFrame(dashboard_data)
dashboard_df = dashboard_df.sort_values(['category', 'z_score'], ascending=[True, False])

# Write to dashboard
current_category = None
for _, data in dashboard_df.iterrows():
    # Add category header if new category
    if data['category'] != current_category:
        dashboard[f'A{row}'] = data['category'].upper()
        dashboard[f'A{row}'].font = category_font
        dashboard[f'A{row}'].fill = category_fill
        dashboard.merge_cells(f'A{row}:J{row}')
        row += 1
        current_category = data['category']

    # Write data
    dashboard[f'A{row}'] = data['code']
    dashboard[f'B{row}'] = data['name']
    dashboard[f'C{row}'] = data['category']
    dashboard[f'D{row}'] = round(data['current_pct'], 1)
    dashboard[f'E{row}'] = round(data['z_score'], 2)
    dashboard[f'F{row}'] = round(data['percentile'], 1)
    dashboard[f'G{row}'] = data['status']
    dashboard[f'H{row}'] = round(data['min_pct'], 1)
    dashboard[f'I{row}'] = round(data['max_pct'], 1)
    dashboard[f'J{row}'] = data['latest_date'].strftime('%Y-%m-%d')

    # Format numbers
    dashboard[f'D{row}'].number_format = '0.0"%"'
    dashboard[f'E{row}'].number_format = '0.00'
    dashboard[f'F{row}'].number_format = '0.0"%"'
    dashboard[f'H{row}'].number_format = '0.0"%"'
    dashboard[f'I{row}'].number_format = '0.0"%"'

    # Color code status
    if data['status'] == "EXTREME LONG":
        dashboard[f'G{row}'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        dashboard[f'G{row}'].font = Font(bold=True, color="FFFFFF")
    elif data['status'] == "EXTREME SHORT":
        dashboard[f'G{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        dashboard[f'G{row}'].font = Font(bold=True, color="FFFFFF")
    elif data['status'] == "Strong Long":
        dashboard[f'G{row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    elif data['status'] == "Strong Short":
        dashboard[f'G{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Apply borders
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        dashboard[f'{col}{row}'].border = border

    row += 1

# Create individual contract sheets with charts
print("\nCreating individual contract sheets with charts...")
for code, data in contract_data.items():
    print(f"  ✓ Creating sheet for {code}")

    ws = wb.create_sheet(code)

    # Title
    ws['A1'] = f"{data['info']['name']} ({code})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Commercial Net Position as % of Open Interest"
    ws['A2'].font = Font(size=10)

    # Statistics
    ws['A4'] = "Statistics (3-Year Period):"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = f"Mean: {data['stats']['mean']:.1f}%"
    ws['A6'] = f"Std Dev: {data['stats']['std']:.1f}%"
    ws['A7'] = f"Min: {data['stats']['min']:.1f}%"
    ws['A8'] = f"Max: {data['stats']['max']:.1f}%"

    # Write data starting at row 10
    ws['A10'] = "Date"
    ws['B10'] = "Commercial Net %"
    ws['A10'].font = header_font
    ws['B10'].font = header_font
    ws['A10'].fill = header_fill
    ws['B10'].fill = header_fill

    # Write historical data
    for idx, (_, row_data) in enumerate(data['df'].iterrows(), start=11):
        ws[f'A{idx}'] = row_data['Date']
        ws[f'B{idx}'] = round(row_data['Commercial_Net_Pct'], 2)
        ws[f'A{idx}'].number_format = 'yyyy-mm-dd'
        ws[f'B{idx}'].number_format = '0.00'

    # Create chart
    chart = LineChart()
    chart.title = f"{data['info']['name']} - Commercial Net Position % OI"
    chart.style = 2
    chart.y_axis.title = "Net Position (% of OI)"
    chart.x_axis.title = "Date"
    chart.height = 10
    chart.width = 20

    # Add data to chart
    data_ref = Reference(ws, min_col=2, min_row=10, max_row=10+len(data['df']))
    dates_ref = Reference(ws, min_col=1, min_row=11, max_row=10+len(data['df']))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(dates_ref)

    # Add chart to sheet
    ws.add_chart(chart, "D4")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15

# Save workbook
output_file = 'tools/CoT_Positioning_Report.xlsx'
wb.save(output_file)

print("\n" + "="*80)
print("EXCEL REPORT CREATED SUCCESSFULLY!")
print("="*80)
print(f"File: {output_file}")
print(f"Sheets: 1 Dashboard + {len(contract_data)} Individual Contracts")
print("\nFeatures:")
print("  ✓ Main dashboard with all contracts by category")
print("  ✓ Z-scores and percentiles vs 3-year history")
print("  ✓ Color-coded extreme positioning alerts")
print("  ✓ Individual sheets with historical charts for each contract")
print("  ✓ Configurable lookback period (Dashboard!B4)")
print("="*80)

"""
Add VBA macro and update button to existing Excel report
This creates a macro-enabled version (.xlsm) with an update button
"""
import os
import sys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

print("="*80)
print("ADDING VBA MACRO TO COT REPORT")
print("="*80)

# File paths
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
INPUT_FILE = os.path.join(PROJECT_ROOT, 'tools', 'CoT_Positioning_Report.xlsx')
OUTPUT_FILE = os.path.join(PROJECT_ROOT, 'tools', 'CoT_Positioning_Report.xlsm')

if not os.path.exists(INPUT_FILE):
    print(f"❌ Excel file not found: {INPUT_FILE}")
    sys.exit(1)

print(f"\nLoading workbook: {INPUT_FILE}")
wb = load_workbook(INPUT_FILE)

# VBA code for the update macro
vba_code = '''Private Sub Workbook_Open()
    ' Clear log on new session
    On Error Resume Next
    Dim logSheet As Worksheet
    Set logSheet = ThisWorkbook.Sheets("UpdateLog")
    If Not logSheet Is Nothing Then
        logSheet.Cells.Clear
        logSheet.Range("A1").Value = "Session Started: " & Now()
        logSheet.Range("A2").Value = String(80, "=")
    End If
    On Error GoTo 0
End Sub

Private Sub LogMessage(msg As String)
    ' Log message to UpdateLog sheet with timestamp
    On Error Resume Next
    Dim logSheet As Worksheet
    Set logSheet = ThisWorkbook.Sheets("UpdateLog")

    If logSheet Is Nothing Then
        ' Create log sheet if it doesn't exist
        Set logSheet = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
        logSheet.Name = "UpdateLog"
        logSheet.Tab.Color = RGB(255, 255, 0) ' Yellow tab
    End If

    Dim nextRow As Long
    nextRow = logSheet.Cells(logSheet.Rows.Count, 1).End(xlUp).Row + 1

    logSheet.Cells(nextRow, 1).Value = Now()
    logSheet.Cells(nextRow, 2).Value = msg

    ' Auto-fit columns
    logSheet.Columns("A:B").AutoFit

    On Error GoTo 0
End Sub

Sub UpdateCoTReport()
    '
    ' Update CoT Report Macro
    ' Reads lookback period from Dashboard!B4 and regenerates the report
    ' Keeps Excel open and automatically reopens the updated file
    '

    Dim lookbackYears As Integer
    Dim projectPath As String
    Dim pythonScript As String
    Dim shellCommand As String
    Dim ws As Worksheet
    Dim wsh As Object
    Dim workbookPath As String
    Dim returnCode As Integer

    On Error GoTo ErrorHandler

    LogMessage "=== UPDATE STARTED ==="
    LogMessage "Button clicked by user"

    ' Disable screen updating for better UX
    Application.ScreenUpdating = False
    Application.DisplayAlerts = False

    ' Get lookback period from Dashboard
    LogMessage "Reading lookback period from Dashboard!B4"
    Set ws = ThisWorkbook.Sheets("Dashboard")
    On Error Resume Next
    lookbackYears = ws.Range("B4").Value
    On Error GoTo ErrorHandler

    LogMessage "Lookback period: " & lookbackYears & " years"

    ' Validate lookback period
    If lookbackYears < 1 Or lookbackYears > 10 Then
        LogMessage "ERROR: Invalid lookback period: " & lookbackYears
        MsgBox "Invalid lookback period in cell B4!" & vbCrLf & _
               "Please enter a value between 1 and 10 years.", _
               vbExclamation, "Invalid Input"
        Application.ScreenUpdating = True
        Application.DisplayAlerts = True
        Exit Sub
    End If

    LogMessage "Validation passed"

    ' Confirm with user
    Dim msg As String
    msg = "This will update the report with a " & lookbackYears & "-year lookback period." & vbCrLf & vbCrLf & _
          "This may take 1-2 minutes. Please wait..." & vbCrLf & vbCrLf & _
          "The file will close and reopen automatically when done." & vbCrLf & vbCrLf & _
          "Continue?"

    LogMessage "Waiting for user confirmation"
    Dim answer As VbMsgBoxResult
    answer = MsgBox(msg, vbQuestion + vbYesNo, "Update CoT Report")

    If answer = vbNo Then
        LogMessage "User cancelled update"
        Application.ScreenUpdating = True
        Application.DisplayAlerts = True
        Exit Sub
    End If

    LogMessage "User confirmed - proceeding with update"

    ' Store the workbook path before closing
    workbookPath = ThisWorkbook.FullName
    LogMessage "Workbook path: " & workbookPath

    ' Get project root path (workbook is in tools/ folder)
    projectPath = ThisWorkbook.Path
    projectPath = Left(projectPath, InStrRev(projectPath, Application.PathSeparator) - 1)
    LogMessage "Project root: " & projectPath

    ' Path to Python script
    pythonScript = projectPath & Application.PathSeparator & "src" & Application.PathSeparator & _
                   "analysis" & Application.PathSeparator & "cot_positioning" & Application.PathSeparator & _
                   "report_generator.py"
    LogMessage "Python script: " & pythonScript

    ' Build shell command with output redirection to capture Python output
    Dim logFilePath As String
    logFilePath = projectPath & Application.PathSeparator & "tools" & Application.PathSeparator & "python_output.log"

    shellCommand = "cmd.exe /c ""cd /d """ & projectPath & """ && " & _
                   "python """ & pythonScript & """ --lookback " & lookbackYears & " > """ & logFilePath & """ 2>&1"""
    LogMessage "Command: " & shellCommand
    LogMessage "Python output will be saved to: " & logFilePath

    ' Save and close the workbook before regenerating
    LogMessage "Saving workbook"
    ThisWorkbook.Save
    LogMessage "Closing workbook"
    ThisWorkbook.Close SaveChanges:=False

    ' Create WScript.Shell object to run command and wait
    LogMessage "Creating WScript.Shell object"
    Set wsh = CreateObject("WScript.Shell")

    ' Run the Python script and wait for completion
    ' 0 = hidden window, True = wait for completion
    LogMessage "Executing Python script (this may take 1-2 minutes)"
    returnCode = wsh.Run(shellCommand, 0, True)
    LogMessage "Python script completed with return code: " & returnCode

    ' Read Python output log and add to UpdateLog
    LogMessage "--- PYTHON OUTPUT START ---"
    On Error Resume Next
    Dim fso As Object, pythonLog As Object, line As String
    Set fso = CreateObject("Scripting.FileSystemObject")
    If fso.FileExists(logFilePath) Then
        Set pythonLog = fso.OpenTextFile(logFilePath, 1)  ' 1 = ForReading
        Do While Not pythonLog.AtEndOfStream
            line = pythonLog.ReadLine
            LogMessage line
        Loop
        pythonLog.Close
    Else
        LogMessage "ERROR: Python output log file not found!"
    End If
    On Error GoTo ErrorHandler
    LogMessage "--- PYTHON OUTPUT END ---"

    ' Check if script ran successfully
    If returnCode = 0 Then
        LogMessage "SUCCESS: Report generation completed"
        ' Reopen the updated workbook
        LogMessage "Reopening workbook: " & workbookPath
        Application.Workbooks.Open workbookPath
        LogMessage "Workbook reopened successfully"
        MsgBox "Report updated successfully!" & vbCrLf & vbCrLf & _
               "Lookback period: " & lookbackYears & " years" & vbCrLf & vbCrLf & _
               "Check the UpdateLog sheet for details.", _
               vbInformation, "Update Complete"
    Else
        LogMessage "ERROR: Python script failed with code " & returnCode
        MsgBox "Update failed! Return code: " & returnCode & vbCrLf & vbCrLf & _
               "Please check if Python is installed and in PATH." & vbCrLf & vbCrLf & _
               "See UpdateLog sheet for details.", _
               vbCritical, "Update Failed"
    End If

    ' Clean up
    Set wsh = Nothing
    Application.ScreenUpdating = True
    Application.DisplayAlerts = True

    LogMessage "=== UPDATE COMPLETED ==="
    Exit Sub

ErrorHandler:
    LogMessage "FATAL ERROR: " & Err.Description & " (Error " & Err.Number & ")"
    MsgBox "An error occurred: " & Err.Description & vbCrLf & vbCrLf & _
           "See UpdateLog sheet for details.", _
           vbCritical, "Error"
    Application.ScreenUpdating = True
    Application.DisplayAlerts = True

End Sub
'''

# Add VBA module to workbook
# Note: openpyxl has limited VBA support, we need to manually add it
print("\n⚠ Note: openpyxl cannot directly add VBA code to workbooks.")
print("Creating instruction file for manual VBA setup...")

# Save VBA code to a separate file
vba_file = os.path.join(PROJECT_ROOT, 'tools', 'Excel_VBA_Code.txt')
with open(vba_file, 'w') as f:
    f.write("VBA CODE FOR UPDATE BUTTON\n")
    f.write("="*80 + "\n\n")
    f.write("INSTRUCTIONS:\n")
    f.write("1. Open CoT_Positioning_Report.xlsx\n")
    f.write("2. Create a new sheet named 'UpdateLog' (right-click sheet tabs > Insert > Worksheet)\n")
    f.write("   - Rename it to 'UpdateLog'\n")
    f.write("   - Optional: Right-click tab > Tab Color > Yellow\n")
    f.write("3. Press Alt+F11 to open VBA Editor\n")
    f.write("4. In the VBA Project Explorer, double-click 'ThisWorkbook'\n")
    f.write("5. Copy the Workbook_Open and LogMessage code (first 2 functions below)\n")
    f.write("6. Insert > Module (to create a new module)\n")
    f.write("7. Copy the UpdateCoTReport code (last function below)\n")
    f.write("8. Close VBA Editor (Alt+Q)\n")
    f.write("9. Go to Dashboard sheet\n")
    f.write("10. Developer tab > Insert > Button (Form Control)\n")
    f.write("11. Draw button in a visible location (e.g., above the data table)\n")
    f.write("12. In 'Assign Macro' dialog, select 'UpdateCoTReport'\n")
    f.write("13. Right-click button > Edit Text > Change to 'Update Report'\n")
    f.write("14. Save as Macro-Enabled Workbook: File > Save As > Excel Macro-Enabled Workbook (.xlsm)\n")
    f.write("    Save to: tools/CoT_Positioning_Report.xlsm\n\n")
    f.write("NOTE: The Workbook_Open and LogMessage subs must go in 'ThisWorkbook'\n")
    f.write("      The UpdateCoTReport sub goes in a Module\n\n")
    f.write("="*80 + "\n")
    f.write("VBA CODE:\n")
    f.write("="*80 + "\n\n")
    f.write(vba_code)
    f.write("\n\n")
    f.write("="*80 + "\n")
    f.write("USAGE:\n")
    f.write("="*80 + "\n\n")
    f.write("After setting up the macro:\n")
    f.write("1. Change lookback period in cell B4 (Dashboard sheet)\n")
    f.write("2. Click the 'Update Report' button\n")
    f.write("3. Confirm the update\n")
    f.write("4. Wait while Python script runs (file closes temporarily)\n")
    f.write("5. File automatically reopens with updated data!\n")
    f.write("6. Check UpdateLog sheet to see what happened\n\n")
    f.write("LOGGING FEATURES:\n")
    f.write("- UpdateLog sheet (yellow tab) tracks each update attempt\n")
    f.write("- Logs are cleared automatically when you open the workbook (new session)\n")
    f.write("- If update fails, check UpdateLog for error details\n")
    f.write("- Each log entry has a timestamp\n\n")
    f.write("NOTE: Excel stays open during the update. Only the workbook closes and reopens.\n\n")
    f.write("NOTE: If you don't want to use the button, you can still use the batch files:\n")
    f.write("- Windows: Double-click tools/update_cot_report.bat\n")
    f.write("- Linux/Mac: Run ./tools/update_cot_report.sh\n")

print(f"\n✓ VBA code saved to: {vba_file}")
print("\n" + "="*80)
print("MANUAL SETUP REQUIRED")
print("="*80)
print("\nSince openpyxl cannot directly add VBA macros, please follow these steps:")
print(f"\n1. Open: {vba_file}")
print("2. Follow the instructions to add the VBA code manually")
print("3. Save as .xlsm file")
print("\nAlternatively, keep using the batch files (no Excel macro needed):")
print("  - Windows: Double-click tools/update_cot_report.bat")
print("  - Linux/Mac: Run ./tools/update_cot_report.sh")
print("="*80)
